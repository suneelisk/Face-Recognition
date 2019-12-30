import face_recognition
import cv2
import numpy as np
import pandas as pd
import time
import threading
from datetime import datetime
from datetime import date
import openpyxl
from tkinter import *
import smtplib
from tkinter import filedialog
from tkinter import *
import os
from PIL import Image, ImageTk

def output_button():
    global folder_path
    filename = filedialog.askdsirectory()
    folder_path.set(filename)
    return filename

def browse_button():
    # Allow user to select a directory and store it in global var
    # called folder_path
    global folder_path
    filename = filedialog.askdirectory()
    folder_path.set(filename)
    print(filename)


    video_capture = cv2.VideoCapture(0)

    data = pd.read_csv(filename+'\data.csv')
    #print(data)

    asa = []

    for img in data['Image']:
        image = face_recognition.load_image_file(img)
        img1 = img.replace('.','')
        locals()["Name_" + str(img1)] = face_recognition.face_encodings(image)[0]
        asa.append(locals()["Name_" + str(img1)])



    # Create arrays of known face encodings and their name
    known_face_encodings =  asa
    known_face_names = list(data["Details"])
    #print(known_face_encodings)


    main_loop_running = True
    def capture():
        global main_loop_running, frame
        while (main_loop_running):
            ret, frame = video_capture.read()
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = small_frame[:, :, ::-1]
            
            if main_loop_running:
                face_locations = face_recognition.face_locations(rgb_small_frame)
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
                
            face_names = []
            
            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Unknown"

                # # If a match was found in known_face_encodings, just use the first one.
                # if True in matches:
                #     first_match_index = matches.index(True)
                #     name = known_face_names[first_match_index]
                
                # Or instead, use the known face with the smallest distance to the new face
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                    
                face_names.append(name)
                now = datetime.now()
                dates = date.today()
                times = datetime.time(datetime.now())
                
                print(name, "Loggedin", now)
                
                

                wb = openpyxl.load_workbook(r'D:\suneel\Nokia Employee Access\output.xlsx')
                ws = wb.active
                row = ws.max_row
                
                ws['A'+str(row+1)] = name
                ws['B'+str(row+1)] = dates
                ws['C'+str(row+1)] = times
                
                wb.save(r'D:\suneel\Nokia Employee Access\output.xlsx')
                
                def check(list1, val):
                    for x in list1:
                        
                        if val == x:
                            return True
                    else:
                        return False
     

                
                
                if check(data['Details'], name):
                    #print(i)
                    root = Tk()
                    root.title("Faec Recognition")
                    root.geometry('300x200')
                    root.configure(background="black")
                    
                    message = "Success Fully Logged in"+str(' ')+str(name)
                    
                    messagevar = Message(root, text = message, width = 250)
                    messagevar.config(bg = 'white',font=("Courier", 10), fg = 'green')
                    messagevar.pack(side="top", padx = 25, pady = 50)
                    root.after(2500, lambda:root.destroy())
                    root.mainloop()
                else:
                    '''gmail_user = "Sender mail id"
                    gmail_pwd = "password"
                    TO = 'Receiver mail id'
                    SUBJECT = "Testing sending using gmail"
                    TEXT = "Unknown person tried to access your Machine"
                    server = smtplib.SMTP('smtp.gmail.com', 587)
                    server.ehlo()
                    server.starttls()
                    server.login(gmail_user, gmail_pwd)
                    BODY = '\r\n'.join(['To: %s' % TO,
                            'From: %s' % gmail_user,
                            'Subject: %s' % SUBJECT,
                            '', TEXT])
                    
                    server.sendmail(gmail_user, [TO], BODY)'''
                    
                    root1 = Tk()
                    root1.title('Face Not Recognized')
                    root1.geometry('300x200')
                    root1.configure(background = 'black')
                    
                    message1 = "Your face was not recognized please wait"
                    
                    messagevar1 = Message(root1, text = message1, width = 250)
                    messagevar1.config(bg = 'white', font = ("Courier", 10), fg = 'red')
                    messagevar1.pack(side = 'top', padx = 25, pady = 50)
                    root1.after(2500, lambda: root1.destroy())
                    root1.mainloop()
                    
                    
                        
                        
            time.sleep(5)

    ret, frame = video_capture.read()
    cv2.imshow('Webcam', frame)
    child_t = threading.Thread(target=capture)
    child_t.setDaemon(True)
    child_t.start()

    while(1):
        ret, frame = video_capture.read()
        cv2.imshow('Webcam', frame)

        # here I want to call capture() function every 3 seconds

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    main_loop_running = False
    child_t.join()

    video_capture.release()
    cv2.destroyAllWindows()


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

Logo = resource_path(r"D:\suneel\Nokia Employee Access\logo.jpg")
Logo1 = resource_path(r"D:\suneel\Nokia Employee Access\logo1.png")

## Main Panel
root = Tk()
root.title("Converting Voice to Text")
root.geometry('500x400')
root.configure(background="black")


class Example(Frame):
    def __init__(self, master, *pargs):
        Frame.__init__(self, master, *pargs)
 
        self.image = Image.open(Logo)
        self.img_copy= self.image.copy()
 
 
        self.background_image = ImageTk.PhotoImage(self.image, format="gif -index 2")
 
        self.background = Label(self, image=self.background_image)
        self.background.pack(fill=BOTH, expand=YES)
        self.background.bind('<Configure>', self._resize_image)
 
        mi_but1=ImageTk.PhotoImage(Image.open(Logo1))
        button2 = Button(self.background, text='button2',image=mi_but1)
        button2.image=mi_but1
        button2.pack(side='top')
        
        link1 = Label(self.background, text="SeaportAI Analytics | Robotics", fg="blue", cursor="hand2")
        link1.pack(side="top", padx = 15, pady = 15)
        link1.config(width=25, height = 2)
        link1.bind("<Button-1>", lambda e: callback("https://seaportai.com/"))
        
        folder_path = StringVar()
        #button1 = Button(text="Output Dictionary", command=output_button, bg="orange", fg="white")
        #button1.pack(side="bottom")
        
        folder_path1 = StringVar()
        button3 = Button(self.background, text="Input Files Dictionary", command=browse_button, bg="orange", fg="black", height = 2, width = 25)
        button3.pack(side = "top", padx = 15, pady = 15)
        
        link2 = Label(self.background, text="Need Help? Contact Us", fg="blue", cursor="hand2")
        link2.pack(side = "top", padx = 15, pady = 15)
        link2.config(width=25, height = 2)
        link2.bind("<Button-1>", lambda e: callback("https://seaportai.com/contact"))
 
    def _resize_image(self,event):
 
        new_width = event.width
        new_height = event.height
 
        self.image = self.img_copy.resize((new_width, new_height))
 
        self.background_image = ImageTk.PhotoImage(self.image)
        self.background.configure(image =  self.background_image)
        





#lbl = Label(root, text="SeaportAI Analytics | Robotics")
#lbl.grid(column = 4, row = 2, padx=15, pady=15)

folder_path = StringVar()
#lbl1 = Label(master=root,textvariable=folder_path)
#lbl1.grid(row=26, column=5)




#lbl11 = Label(master=root,textvariable=folder_path1)
#lbl11.grid(row=5, column=1)



e = Example(root)
e.pack(fill=BOTH, expand=YES)
root.mainloop()

